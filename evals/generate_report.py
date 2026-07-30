"""Generate XLS evaluation reports with summary + detail sheets.

Called by run_routing_eval.py and run_response_eval.py after eval runs.
Can also be run standalone to regenerate XLS from existing CSV/JSON reports.

Usage:
    python -m evals.generate_report                          # regen latest
    python -m evals.generate_report --csv reports/routing-eval-*.csv
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPORTS_DIR = Path(__file__).resolve().parent / "reports"

_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BOLD = Font(bold=True, size=11)
_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        widths = []
        for cell in col:
            if cell.value:
                widths.append(len(str(cell.value)))
        width = min(max(widths + [min_width]), max_width)
        ws.column_dimensions[col_letter].width = width + 2


def _apply_verdict_colors(ws, col_idx: int, start_row: int = 2):
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = str(cell.value).upper() if cell.value else ""
        if val == "PASS":
            cell.fill = _GREEN
        elif val == "PASS_CATEGORY":
            cell.fill = _YELLOW
        elif val == "FAIL":
            cell.fill = _RED


def generate_xls_from_routing(
    results: list[dict], summary: dict, ts: str
) -> Path:
    """Generate XLS workbook for routing eval."""
    wb = Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    summary_rows = [
        ("Metric", "Value"),
        ("Total cases", summary["total"]),
        ("Exact pass", f"{summary['exact_pass']} ({summary['exact_rate']}%)"),
        ("Category pass", summary["category_pass"]),
        ("Fail", summary["fail"]),
        ("Effective rate", f"{summary['effective_rate']}%"),
        ("Avg latency", f"{summary['avg_latency_ms']}ms"),
        ("", ""),
        ("By Route Type", ""),
    ]
    for rt, stats in summary.get("by_route_type", {}).items():
        rate = round(stats["pass"] / stats["total"] * 100, 1) if stats["total"] else 0
        summary_rows.append((f"  {rt}", f"{stats['pass']}/{stats['total']} ({rate}%)"))

    summary_rows.append(("", ""))
    summary_rows.append(("By Category", ""))
    for cat, stats in summary.get("by_category", {}).items():
        rate = round(stats["pass"] / stats["total"] * 100, 1) if stats["total"] else 0
        summary_rows.append((f"  {cat}", f"{stats['pass']}/{stats['total']} ({rate}%)"))

    for row_data in summary_rows:
        ws_sum.append(row_data)

    ws_sum.cell(row=1, column=1).font = _BOLD
    ws_sum.cell(row=1, column=2).font = _BOLD
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 30

    # --- Detail sheet ---
    ws_det = wb.create_sheet("Details")
    headers = ["question", "expected_slug", "actual_slug", "route_type",
               "category", "pass", "elapsed_ms", "reason"]
    ws_det.append(headers)
    _style_header(ws_det, len(headers))

    for r in results:
        ws_det.append([r.get(h, "") for h in headers])

    pass_col = headers.index("pass") + 1
    _apply_verdict_colors(ws_det, pass_col)
    _auto_width(ws_det)

    # --- Failures sheet ---
    if summary.get("failures"):
        ws_fail = wb.create_sheet("Failures")
        fail_headers = ["question", "expected", "actual"]
        ws_fail.append(fail_headers)
        _style_header(ws_fail, len(fail_headers))
        for f_ in summary["failures"]:
            ws_fail.append([f_["question"], f_["expected"], f_["actual"]])
        _auto_width(ws_fail)

    out = REPORTS_DIR / f"routing-eval-{ts}.xlsx"
    wb.save(out)
    return out


def generate_xls_from_response(
    results: list[dict], summary: dict, ts: str
) -> Path:
    """Generate XLS workbook for response eval."""
    wb = Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    summary_rows = [
        ("Metric", "Value"),
        ("Total cases", summary["total"]),
        ("Pass", f"{summary['pass']} ({summary['pass_rate']}%)"),
        ("Fail", summary["fail"]),
        ("Avg latency", f"{summary['avg_latency_ms']}ms"),
        ("Avg response length", f"{summary['avg_response_length']} chars"),
        ("", ""),
        ("By Category", ""),
    ]
    for cat, stats in summary.get("by_category", {}).items():
        rate = round(stats["pass"] / stats["total"] * 100, 1) if stats["total"] else 0
        summary_rows.append((f"  {cat}", f"{stats['pass']}/{stats['total']} ({rate}%)"))

    for row_data in summary_rows:
        ws_sum.append(row_data)

    ws_sum.cell(row=1, column=1).font = _BOLD
    ws_sum.cell(row=1, column=2).font = _BOLD
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 30

    # --- Detail sheet ---
    ws_det = wb.create_sheet("Details")
    headers = ["expected_slug", "agent_name", "category", "pass",
               "contains_expected", "no_forbidden", "sufficient_length",
               "elapsed_ms", "reason", "response_length", "question",
               "response_preview"]
    ws_det.append(headers)
    _style_header(ws_det, len(headers))

    for r in results:
        ws_det.append([r.get(h, "") for h in headers])

    pass_col = headers.index("pass") + 1
    _apply_verdict_colors(ws_det, pass_col)
    _auto_width(ws_det)

    # --- Per-category sheets ---
    categories = sorted(set(r["category"] for r in results))
    for cat in categories:
        cat_results = [r for r in results if r["category"] == cat]
        ws_cat = wb.create_sheet(cat[:28])
        cat_headers = ["agent_name", "pass", "elapsed_ms", "reason",
                       "response_length", "response_preview"]
        ws_cat.append(cat_headers)
        _style_header(ws_cat, len(cat_headers))
        for r in cat_results:
            ws_cat.append([r.get(h, "") for h in cat_headers])
        _apply_verdict_colors(ws_cat, 2)
        _auto_width(ws_cat)

    # --- Failures sheet ---
    if summary.get("failures"):
        ws_fail = wb.create_sheet("Failures")
        fail_headers = ["slug", "question", "reason"]
        ws_fail.append(fail_headers)
        _style_header(ws_fail, len(fail_headers))
        for f_ in summary["failures"]:
            ws_fail.append([f_["slug"], f_["question"], f_["reason"]])
        _auto_width(ws_fail)

    out = REPORTS_DIR / f"response-eval-{ts}.xlsx"
    wb.save(out)
    return out
