"""FastVC game LLM evaluation — plays through full game scenarios with real LLM.

Tests that the game master (Coach V) produces well-structured, on-topic responses:
  1. Character intro includes Baltic context and 3 numbered choices
  2. Gameplay turns respond to player actions with coaching tone
  3. Game progresses through stages and rounds correctly
  4. Special abilities are acknowledged
  5. Game over produces a scorecard

HITS THE LLM. Expect ~3-5 minutes for all 3 scenarios.

Usage:
    python -m evals.run_game_eval                    # all 3 scenarios
    python -m evals.run_game_eval --scenario 1       # single scenario
    python -m evals.run_game_eval --rounds 2         # fewer rounds per scenario
"""

from __future__ import annotations

import argparse
import json
import re
import time
from datetime import datetime, timezone
from pathlib import Path

from langchain_core.messages import HumanMessage, SystemMessage

from game.engine import (
    CHARACTERS, LEVELS, STAGES, GameState,
    new_game, draw_event, format_status, calculate_score,
)
from game.prompts import GAME_MASTER_SYSTEM
from utils.llm import build_llm

REPORTS_DIR = Path(__file__).resolve().parent / "reports"

SCENARIOS = [
    {
        "id": 1,
        "character": "dealmaker",
        "name": "Marcus Drake",
        "level": "associate",
        "actions": [
            "I want to pursue the largest deal in the pipeline",
            "Build a preliminary LBO model on this target",
            "Run commercial due diligence on the management team",
            "Use my special ability Open Door to reach the founder directly",
            "Negotiate a 6x EBITDA entry multiple with earnout",
            "Implement pricing optimization across the portfolio company",
        ],
    },
    {
        "id": 2,
        "character": "analyst",
        "name": "Elena Voss",
        "level": "associate",
        "actions": [
            "2",
            "Use Deep Model to build a forensic LBO on the top target",
            "Flag the revenue quality issue to the IC",
            "Run sensitivity analysis on downside scenarios",
            "Present the investment memo to the committee",
            "Track portfolio company KPIs for quarterly review",
        ],
    },
    {
        "id": 3,
        "character": "investigator",
        "name": "Raj Mehta",
        "level": "associate",
        "actions": [
            "Screen the pipeline for any red flags",
            "Use my Red Flag ability to spot hidden risks",
            "Deep dive into the target's customer concentration",
            "Review change of control clauses in key contracts",
            "Recommend walking away from this deal — too risky",
            "Start sourcing a new, cleaner target",
        ],
    },
]


def _has_numbered_choices(text: str) -> bool:
    """Check if response contains 3 numbered choices (1. 2. 3.)."""
    pattern = re.compile(r"^\s*[1-3]\.\s+", re.MULTILINE)
    matches = pattern.findall(text)
    return len(matches) >= 3


def _has_coaching_tone(text: str) -> bool:
    """Check for coaching/energetic language markers."""
    markers = [
        r"!",
        r"BOOM|GREAT|NICE|SOLID|SMART|ROOKIE|WAKE UP|LET'S|THAT'S",
        r"\*\*",
    ]
    hits = sum(1 for m in markers if re.search(m, text, re.IGNORECASE))
    return hits >= 2


def _has_baltic_context(text: str) -> bool:
    """Check for Baltic market references."""
    baltic_markers = [
        r"Tallinn|Riga|Vilnius|Tartu|Kaunas|Klaipeda",
        r"Estonia|Latvia|Lithuania|Baltic",
        r"€\d",
        r"EBITDA|revenue|multiple",
    ]
    hits = sum(1 for m in baltic_markers if re.search(m, text, re.IGNORECASE))
    return hits >= 2


def _has_pe_content(text: str) -> bool:
    """Check for VC-specific terminology."""
    pe_terms = [
        r"deal|portfolio|fund|invest|exit|return|multiple|leverage",
        r"LBO|EBITDA|IRR|MOIC|EV|valuation|diligence",
        r"LP|GP|carry|capital|deploy|source|close",
    ]
    hits = sum(1 for m in pe_terms if re.search(m, text, re.IGNORECASE))
    return hits >= 2


def _score_response(text: str, turn_type: str) -> dict:
    """Score a single game response."""
    checks = {
        "has_content": len(text) > 100,
        "has_numbered_choices": _has_numbered_choices(text),
        "has_coaching_tone": _has_coaching_tone(text),
        "has_pe_content": _has_pe_content(text),
        "no_refusal": not re.search(
            r"I can't|I cannot|I'm not able|as an AI|I apologize",
            text, re.IGNORECASE,
        ),
    }

    if turn_type == "intro":
        checks["has_baltic_context"] = _has_baltic_context(text)
        checks["has_company_names"] = bool(re.search(r"\*\*[A-Z]", text))

    passed = all(checks.values())
    return {"pass": passed, "checks": checks}


def _build_system(state: GameState) -> str:
    char = CHARACTERS.get(state.character, {})
    lvl = LEVELS.get(state.level, {})
    event = draw_event()
    state.events_history.append(event["name"])

    char_info = (
        f"**{char['name']}** — {char['title']} ({char['icon']})\n"
        f"Role: {char['role']}\n"
        f"Ability: {char['ability']}\n"
        f"Background: {char['description']}"
    )

    return GAME_MASTER_SYSTEM.format(
        total_rounds=state.total_rounds,
        status=format_status(state),
        event=f"**{event['name']}**: {event['effect']}",
        character_info=char_info,
        level_title=lvl.get("title", "Associate"),
        level_complexity=lvl.get("complexity", ""),
    )


def run_scenario(scenario: dict, max_rounds: int = 3) -> dict:
    """Play through a scenario and score each turn."""
    char_key = scenario["character"]
    state = new_game(char_key)
    llm = build_llm()
    results = []
    actions = scenario["actions"][:max_rounds * 2]

    print(f"\n  Scenario {scenario['id']}: {scenario['name']} ({char_key})")

    # Turn 1: Character intro
    system = _build_system(state)
    messages = [
        SystemMessage(content=system),
        HumanMessage(content=(
            "The game begins! Present Round 1, Stage 1: Deal Sourcing.\n"
            "Set the scene — the player just joined a Baltic VC fund. "
            "Show 3-4 potential deals in the pipeline with company names, countries, sectors, revenues.\n"
            "Give your coaching intro — fire them up! Then end with 3 choices."
        )),
    ]

    print(f"    [1/{len(actions)+1}] intro...", end=" ", flush=True)
    t0 = time.time()
    response = llm.invoke(messages)
    intro_text = response.content if hasattr(response, "content") else str(response)
    elapsed = round((time.time() - t0) * 1000)

    score = _score_response(intro_text, "intro")
    verdict = "PASS" if score["pass"] else "FAIL"
    print(f"{verdict} ({elapsed}ms, {len(intro_text)} chars)")

    results.append({
        "turn": 0,
        "turn_type": "intro",
        "action": "(character select)",
        "pass": verdict,
        "checks": score["checks"],
        "elapsed_ms": elapsed,
        "response_length": len(intro_text),
        "response_preview": intro_text[:200].replace("\n", " "),
    })

    # Gameplay turns
    for i, action in enumerate(actions):
        system = _build_system(state)
        messages = [
            SystemMessage(content=system),
            HumanMessage(content=(
                f"Player action: {action}\n\n"
                f"Process this for {state.current_stage()} (Round {state.round}/{state.total_rounds}).\n"
                f"React to their choice — give coaching feedback (praise great moves, roast bad ones).\n"
                f"Show the outcome with updated resource numbers.\n"
                f"Then present 3 new choices for the next action."
            )),
        ]

        print(f"    [{i+2}/{len(actions)+1}] {action[:40]}...", end=" ", flush=True)
        t0 = time.time()
        response = llm.invoke(messages)
        turn_text = response.content if hasattr(response, "content") else str(response)
        elapsed = round((time.time() - t0) * 1000)

        turn_type = "special" if "special" in action.lower() or "ability" in action.lower() else "action"
        score = _score_response(turn_text, turn_type)
        verdict = "PASS" if score["pass"] else "FAIL"
        print(f"{verdict} ({elapsed}ms, {len(turn_text)} chars)")

        results.append({
            "turn": i + 1,
            "turn_type": turn_type,
            "action": action,
            "pass": verdict,
            "checks": score["checks"],
            "elapsed_ms": elapsed,
            "response_length": len(turn_text),
            "response_preview": turn_text[:200].replace("\n", " "),
        })

        # Advance state
        if any(kw in turn_text.lower() for kw in ["next stage", "stage complete", "moving to"]):
            state.stage_idx = min(state.stage_idx + 1, len(STAGES) - 1)

    return {
        "scenario_id": scenario["id"],
        "character": scenario["name"],
        "character_key": char_key,
        "total_turns": len(results),
        "passed": sum(1 for r in results if r["pass"] == "PASS"),
        "failed": sum(1 for r in results if r["pass"] == "FAIL"),
        "avg_latency_ms": round(sum(r["elapsed_ms"] for r in results) / len(results)),
        "avg_response_length": round(sum(r["response_length"] for r in results) / len(results)),
        "turns": results,
    }


def main():
    parser = argparse.ArgumentParser(description="Run FastVC game LLM evaluation")
    parser.add_argument("--scenario", type=int, help="Run single scenario (1-3)")
    parser.add_argument("--rounds", type=int, default=3, help="Max rounds per scenario")
    args = parser.parse_args()

    scenarios = SCENARIOS
    if args.scenario:
        scenarios = [s for s in SCENARIOS if s["id"] == args.scenario]

    print(f"Running {len(scenarios)} game scenario(s) with {args.rounds} max rounds (hits LLM)...")

    all_results = []
    for scenario in scenarios:
        result = run_scenario(scenario, max_rounds=args.rounds)
        all_results.append(result)

    # Summary
    total_turns = sum(r["total_turns"] for r in all_results)
    total_pass = sum(r["passed"] for r in all_results)
    total_fail = sum(r["failed"] for r in all_results)
    pass_rate = round(total_pass / total_turns * 100, 1) if total_turns else 0

    summary = {
        "total_scenarios": len(all_results),
        "total_turns": total_turns,
        "pass": total_pass,
        "fail": total_fail,
        "pass_rate": pass_rate,
        "avg_latency_ms": round(sum(r["avg_latency_ms"] for r in all_results) / len(all_results)),
        "scenarios": all_results,
    }

    # Write reports
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    json_path = REPORTS_DIR / f"game-eval-{ts}.json"
    json_path.write_text(json.dumps(summary, indent=2))

    # CSV flat view
    import csv
    csv_path = REPORTS_DIR / f"game-eval-{ts}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "scenario", "character", "turn", "turn_type", "action",
            "pass", "has_content", "has_numbered_choices", "has_coaching_tone",
            "has_pe_content", "no_refusal", "has_baltic_context",
            "has_company_names", "elapsed_ms", "response_length",
            "response_preview",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for result in all_results:
            for turn in result["turns"]:
                row = {
                    "scenario": result["scenario_id"],
                    "character": result["character"],
                    "turn": turn["turn"],
                    "turn_type": turn["turn_type"],
                    "action": turn["action"],
                    "pass": turn["pass"],
                    "elapsed_ms": turn["elapsed_ms"],
                    "response_length": turn["response_length"],
                    "response_preview": turn["response_preview"],
                }
                row.update({k: v for k, v in turn["checks"].items()})
                w.writerow(row)

    # XLS report
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        bold = Font(bold=True, size=11)

        summary_rows = [
            ("Metric", "Value"),
            ("Total scenarios", summary["total_scenarios"]),
            ("Total turns", summary["total_turns"]),
            ("Pass", f"{summary['pass']} ({summary['pass_rate']}%)"),
            ("Fail", summary["fail"]),
            ("Avg latency", f"{summary['avg_latency_ms']}ms"),
        ]
        for row in summary_rows:
            ws.append(row)
        ws.cell(row=1, column=1).font = bold
        ws.cell(row=1, column=2).font = bold
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

        for result in all_results:
            ws_s = wb.create_sheet(f"Scenario {result['scenario_id']}")
            headers = ["Turn", "Type", "Action", "Pass", "Content", "Choices",
                        "Coaching", "VC Terms", "No Refusal", "Latency (ms)",
                        "Length", "Preview"]
            ws_s.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws_s.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for turn in result["turns"]:
                c = turn["checks"]
                ws_s.append([
                    turn["turn"],
                    turn["turn_type"],
                    turn["action"][:60],
                    turn["pass"],
                    c.get("has_content", ""),
                    c.get("has_numbered_choices", ""),
                    c.get("has_coaching_tone", ""),
                    c.get("has_pe_content", ""),
                    c.get("no_refusal", ""),
                    turn["elapsed_ms"],
                    turn["response_length"],
                    turn["response_preview"][:80],
                ])
                row_idx = ws_s.max_row
                pass_cell = ws_s.cell(row=row_idx, column=4)
                pass_cell.fill = green if turn["pass"] == "PASS" else red

            for col in ws_s.columns:
                from openpyxl.utils import get_column_letter
                letter = get_column_letter(col[0].column)
                max_len = max(len(str(c.value or "")) for c in col)
                ws_s.column_dimensions[letter].width = min(max_len + 2, 50)

        xls_path = REPORTS_DIR / f"game-eval-{ts}.xlsx"
        wb.save(xls_path)
    except Exception as e:
        xls_path = None
        print(f"  (XLS generation failed: {e})")

    print(f"\nResults:")
    print(f"  Scenarios:    {summary['total_scenarios']}")
    print(f"  Total turns:  {summary['total_turns']}")
    print(f"  Pass:         {summary['pass']} ({summary['pass_rate']}%)")
    print(f"  Fail:         {summary['fail']}")
    print(f"  Avg latency:  {summary['avg_latency_ms']}ms")

    print(f"\nReports:")
    print(f"  {csv_path}")
    print(f"  {json_path}")
    if xls_path:
        print(f"  {xls_path}")


if __name__ == "__main__":
    main()
